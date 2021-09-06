import pandas as pd 
from openpyxl import load_workbook
import xlsxwriter
import os.path
from os import path
import datetime 
import PySimpleGUI as sg 

class setUp():

	#attributes
	def __init__(self, workbookName, worksheetName, rate):
		self.rate = rate 
		self.workbookName = workbookName
		self.worksheetName = worksheetName

	#creates the excel in the folder and returns as an object
	def initializeWorkbook(self, workbookName):
		outWorkbook = xlsxwriter.Workbook(workbookName + ".xlsx")
		return outWorkbook

	#loads the previous excel's data into a dictionary, each row is a dictionary key with a list of the row's data as the value (some data is strings).
	#used to re-write
	def read(self, workBookName): 
		previousContents = {"Row 1": []}
		print(self.workbookName)
		wb = load_workbook(self.workbookName + ".xlsx")
		worksheet = wb.active
		for i in range(0, worksheet.max_row):
			for col in worksheet.iter_cols(0, worksheet.max_column): 
				location = "Row "+str(col[i].row)
				if(col[i].value == None):
					pass
				elif(previousContents.get(location) == None):
					previousContents[location] = [col[i].value]
				else:
					#cannot append strings in dictionaries unless the value of the key is a LIST
					previousContents[location].append(col[i].value)

		return previousContents
	def writeNewWeekAndTotal(self, tempExcelWorksheet, tempDictionary):
		excelWorksheet = tempExcelWorksheet
		informationDictionary = tempDictionary

	def rewrite(self, tempExcelWorksheet, tempDictionary):
		excelWorksheet = tempExcelWorksheet
		informationDictionary = tempDictionary

		i = 1
		#organizes scattered data into two lines: hours and date - transcribes the previous date
		keyList = (list(informationDictionary))
		while i < len(keyList):
			
			stringIndexTwo = len(keyList[i])
			#print(keyList)
			holder = stringIndexTwo - 4
			increment = 1
			stringIndices = ""
			while holder > 0:
				#print(informationDictionary[keyList[i]])
				#print(len(informationDictionary[keyList[i]]))
				stringIndices += str(keyList[i][3+increment])
				increment += 1
				holder -= 1
				
			counter = 0	
			while counter < len(informationDictionary[keyList[i]]):
				print(len(informationDictionary[keyList[i]]))
				if counter > 2:
					excelWorksheet.write("G"+stringIndices, informationDictionary[keyList[i]][counter])
					excelWorksheet.write("J"+stringIndices, informationDictionary[keyList[i]][counter])	
					counter += 1
				else:	
					print(informationDictionary[keyList[i]])
					print(counter)
					if(counter == 0):
						print("1")
						excelWorksheet.write("D"+stringIndices, informationDictionary[keyList[i]][counter])
					elif(counter == 1):
						print("2")
						excelWorksheet.write("F"+stringIndices, informationDictionary[keyList[i]][counter])
					else:
						print("3")
						excelWorksheet.write("H"+stringIndices, informationDictionary[keyList[i]][counter])
					
					counter += 1
				
			i += 1
		
		
	def writeNewDateHoursPair(self, tempExcelWorksheet, location, rate, hours, confirmation, dateToday, dateWorked):
		excelWorksheet = tempExcelWorksheet
		locationInt = int(location)
		location = locationInt+1
		insertPlace = str(location)
		print(insertPlace)
		addDates = True  
		dateCut = dateWorked
		print("Yup")
		print(confirmation)
		print("break")
		
		if(confirmation == 'Yes'):
			
			if(dateToday == 'Yes'):
				print("In Here")
				dateInput = datetime.datetime.now()
				dateResult = dateFormat(dateInput)
				hoursWorked = int(hours)	
				excelWorksheet.write("D"+str(insertPlace),  dateResult.returnString())
				excelWorksheet.write("F"+str(insertPlace),  hoursWorked)
				excelWorksheet.write("H"+str(insertPlace), hoursWorked*rate)
				addDates = False
				print(addDates == False)
				
			elif(dateToday =='No'):
				print("Not Here")
				year = int(dateCut[4:8])
				print(year)
				month = int(dateCut[0:2])
				print(month)
				day = int(dateCut[3:4])
				print(day)
				hoursWorked = int(hours)
				dateTime = datetime.datetime(year, month, day)
				dateString = dateFormat(dateTime)
				excelWorksheet.write("D"+str(insertPlace),  dateString.returnString())
				excelWorksheet.write("F"+str(insertPlace),  hoursWorked)
				excelWorksheet.write("H"+str(insertPlace), hoursWorked*rate)
				addDates = False

		else:
			pass

class dateFormat():

	def __init__(self, date):
		self.year = date.year
		self.month = date.month
		self.day = date.day
		self.time = date

	def returnYear(self):
		return str(self.year)

	def returnMonth(self):
		return str(self.month) 

	def returnDay(self):
		return str(self.day)
		
	def returnString(self):
		stringDate = str(self.year) + "/" + str(self.month) + "/" + str(self.day)
		return stringDate
	
def main():
	sg.theme('DarkGreen6')

	#layout = [[sg.popup_get_file("Is this the file you wish to access?")]]
	'''layout = [
	[sg.Text("Hello, hope you had a good shift! Please input your hours below:")], 
	[sg.Input(key = "Hours")], 
	[sg.Text("Was your shift today, or another day?")],
	[sg.OK()],
	[sg.OK()]]'''

	layoutTest = [[sg.FileBrowse(key = "Excel File")],
	[sg.Text("Please put in your rate: ")],
	[sg.Input(key = "Rate")],
	[sg.Text("Please specify if you have dates to add? (Yes/No)")],
	[sg.Input(key = "Confirmation")],
	[sg.Text("Hello, hope you had a good shift! Please input your hours below:")],
	[sg.Input(key = "Hours")],
	[sg.Text("Please specify if you shift was today or not? (Yes/No)")],
	[sg.Input(key = "DateToday")],
	[sg.Text("If you said no, please input the date you worked. (MMDDYYYY)")],
	[sg.Input(key = "DateWorked")],
	[sg.Exit("Save"), sg.Exit("Create")]]

	#Exit event returns correct value
	window = sg.Window('PyHours!', layoutTest, enable_close_attempted_event = True)
	while True:
		event, values = window.read()
		print(event, values)
		estimate = len(values['Excel File'])
		endPoint = estimate-6
		

		

		if event == 'Save' or event == 'Create':

			if(path.exists("WorkProject.xlsx") == False):
				print("Trigger1")
				rate = values['Rate']
				print(rate)
				outWorkbook = setUp("WorkProject", "WorkSheet", rate)
				filename = outWorkbook.workbookName
				sheetname = outWorkbook.worksheetName
				excelWorkbook = outWorkbook.initializeWorkbook(filename)
				excelWorksheet = excelWorkbook.add_worksheet(sheetname)
				
				excelWorksheet.write("A1", "Rate:")
				#rate = input("What is your rate? ")
				excelWorksheet.write("B1", rate)
				excelWorksheet.write("D1", "Date:")
				excelWorksheet.write("F1", "Hours:")
				excelWorksheet.write("H1", "Earned: ")
				excelWorksheet.write("J1", "Weekly Total: ")
				
				excelWorkbook.close()
				break

			elif(path.exists("WorkProject.xlsx") and values['Excel File'] != ''):
				print("Trigger2")
				startPoint = (values['Excel File'][endPoint])
				startPointCounter = endPoint

				while startPoint != '/':
					startPointCounter -= 1
					startPoint = values['Excel File'][startPointCounter]
					fileName = values['Excel File'][startPointCounter+1 : endPoint+1]

				rate = values['Rate']
				correctRate = float(rate)
				#should ask for name to be mutable 
				outWorkbook = setUp(fileName, "WorkSheet", correctRate)
				filename = outWorkbook.workbookName
				sheetname = outWorkbook.worksheetName
				excelWorkbook = outWorkbook.initializeWorkbook(filename)
				excelWorksheet = excelWorkbook.add_worksheet(sheetname)

				#this dictionary holds the contents of the previous excel file.
				informationDictionary = outWorkbook.read(excelWorkbook)

				excelWorksheet.write("A1", "Rate:")
				#rate = input("What is your rate? ")
				excelWorksheet.write("B1", rate)
				excelWorksheet.write("D1", "Date:")
				excelWorksheet.write("F1", "Hours:")
				excelWorksheet.write("H1", "Earned: ")
				excelWorksheet.write("J1", "Weekly Total: ")
				#at this point the work book has the same first line and we have the same data in a dictionary

				keyList = list(informationDictionary.keys())
				lastElement = (keyList[len(keyList)-1])
				print(lastElement)

				lastElementUsefulLength = len(lastElement) - 4
				print(lastElementUsefulLength)
				incrementTool = 1
				lastIndices = ""
				while lastElementUsefulLength > 0:
					lastIndices += str(keyList[len(keyList)-1][3+incrementTool])
					incrementTool += 1
					lastElementUsefulLength -= 1

				print(lastElementUsefulLength)
				outWorkbook.rewrite(excelWorksheet, informationDictionary)
				print('Here')
				outWorkbook.writeNewDateHoursPair(excelWorksheet, lastIndices, correctRate, values['Hours'], values['Confirmation'], values['DateToday'], values['DateWorked'])

				excelWorkbook.close()
				window.close()
			break
		
		elif event == sg.WINDOW_CLOSE_ATTEMPTED_EVENT:
			sg.popup("Please use the 'Save' button")


main()